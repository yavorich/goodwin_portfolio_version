import uuid

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Side, Border
from openpyxl.utils import get_column_letter

from django.conf import settings
from pathlib import Path


class ExcelFileCreator:
    fields = []
    titles = {}
    save_path = ""

    def __init__(self):
        self._wb = Workbook()
        self._sheet = self._wb.active

    def get_fields(self):
        return self.fields

    def to_excel(self, data):
        self._create_title()
        self._write_data(data)
        self._set_dimension()

    def save(self, *args, **kwargs):
        path = self._get_save_path(*args, **kwargs)
        self._wb.save(path)
        return Path(*path.parts[:1], *path.parts[2:])

    def get_save_path(self, *args, **kwargs):
        return self.save_path

    def _get_save_path(self, *args, **kwargs):
        directory = settings.MEDIA_ROOT / "excel" / self.get_save_path(*args, **kwargs)
        Path(directory).mkdir(parents=True, exist_ok=True)
        file_name = f"{uuid.uuid4().hex}.xlsx"
        return directory / file_name

    def _create_title(self):
        bs = Side("thin")
        border = Border(left=bs, right=bs, top=bs, bottom=bs)
        center_alignment = Alignment(horizontal="center")

        for i in range(len(self.fields)):
            self._sheet.cell(row=1, column=i + 1).value = self.titles.get(
                self.fields[i], self.fields[i].capitalize()
            )
            self._sheet.cell(row=1, column=i + 1).alignment = center_alignment
            self._sheet.cell(row=1, column=i + 1).border = border

    def _write_data(self, data):
        for row in data:
            self._sheet.append([row[field] for field in self.fields])

    def _set_dimension(self):
        for column_cells in self._sheet.columns:
            length = max(
                len(str(cell.value)) for cell in column_cells if cell.value is not None
            )
            self._sheet.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = (length + 4)
