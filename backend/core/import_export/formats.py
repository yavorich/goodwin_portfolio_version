from io import BytesIO

from import_export.formats.base_formats import XLSX
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from tablib.formats._xlsx import XLSXFormat, safe_xlsx_sheet_title


class DimensionXLSXFormat(XLSXFormat):
    @classmethod
    def export_set(
        cls, dataset, freeze_panes=True, invalid_char_subst="-", escape=False
    ):
        wb = Workbook()
        ws = wb.worksheets[0]

        ws.title = (
            safe_xlsx_sheet_title(dataset.title, invalid_char_subst)
            if dataset.title
            else "Tablib Dataset"
        )

        cls.dset_sheet(dataset, ws, freeze_panes=freeze_panes, escape=escape)
        cls.set_dimension(ws)

        stream = BytesIO()
        wb.save(stream)
        return stream.getvalue()

    @staticmethod
    def set_dimension(ws):
        for column_cells in ws.columns:
            length = max(
                len(str(cell.value)) for cell in column_cells if cell.value is not None
            )
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = (
                length + 4
            )
